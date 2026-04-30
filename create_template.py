import openpyxl
from openpyxl.styles import Font, Fill, PatternFill

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solar Load Calculator"

    # Define headers
    headers = [
        "Field", "Value", "Calculation", "Result"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Define Input Fields (Rows 2-12)
    inputs = [
        ("Consumer Name", ""),
        ("Consumer Number", ""),
        ("Billing Date", ""),
        ("Billing Period", ""),
        ("Units Consumed (kWh)", 0),
        ("Sanctioned Load (kW)", 0),
        ("Connected Load (kW)", 0),
        ("Tariff Category", ""),
        ("Total Bill Amount", 0),
        ("Due Date", ""),
        ("Meter Number", "")
    ]

    for row_idx, (field, value) in enumerate(inputs, 2):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)

    # Define Formulas (Rows 14+)
    ws.cell(row=14, column=1, value="Calculated Solar Load (kW)")
    ws.cell(row=14, column=3, value="1.2 * Sanctioned Load")
    ws.cell(row=14, column=4, value="=B7 * 1.2") 
    
    ws.cell(row=15, column=1, value="Estimated Monthly Savings")
    ws.cell(row=15, column=3, value="Units * 8")
    ws.cell(row=15, column=4, value="=B6 * 8")

    ws.cell(row=16, column=1, value="ROI Period (Years)")
    ws.cell(row=16, column=3, value="Bill / (Savings * 12)")
    ws.cell(row=16, column=4, value="=ROUND(B10 / (D15 * 12 + 0.001), 2)")

    # Save
    wb.save("assets/template.xlsx")
    print("Template created at assets/template.xlsx")

if __name__ == "__main__":
    import os
    if not os.path.exists("assets"):
        os.makedirs("assets")
    create_template()
