import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import math

def fill_excel_template(data_list, output_path):
    """
    Creates a new Excel file with the specific format requested by the user.
    data_list: List of dictionaries, each containing consumer data and consumption history.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solar Load Calculator"

    # Define styles
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid") # Darker Orange
    label_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Light Peach/Orange
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    # Top Section
    labels = ["Consumer Name", "Consumer No", "Fixed Charges", "Sanct. Load (kW)", "Connection Type"]
    keys = ["consumer_name", "consumer_number", "fixed_charges", "sanctioned_load", "connection_type"]
    
    for i, (label, key) in enumerate(zip(labels, keys), 1):
        cell_label = ws.cell(row=i, column=1, value=label)
        cell_label.font = bold_font
        cell_label.border = border
        # cell_label.fill = label_fill # Optional: based on image

        # Consumer 1 (Column C)
        val1 = data_list[0].get(key, "") if len(data_list) > 0 else ""
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
        cell1 = ws.cell(row=i, column=3, value=val1)
        cell1.alignment = center_align
        for col in range(3, 6): ws.cell(row=i, column=col).border = border
        
        # Consumer 2 (Column G)
        val2 = data_list[1].get(key, "") if len(data_list) > 1 else ""
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=9)
        cell2 = ws.cell(row=i, column=7, value=val2)
        cell2.alignment = center_align
        for col in range(7, 10): ws.cell(row=i, column=col).border = border

    ws.cell(row=6, column=1, value="Contract Demand (KVA) :").font = bold_font
    
    ws.cell(row=7, column=1, value="Solar Pannel used").font = bold_font
    ws.cell(row=7, column=1).border = border
    panel_rating_cell = ws.cell(row=7, column=2, value=600)
    panel_rating_cell.fill = yellow_fill
    panel_rating_cell.border = border
    panel_rating_cell.alignment = center_align

    # Table Header (Row 8)
    headers = ["Sr.No", "Month", "Units", "Bill Amount", "Unit Cost", "Month", "Units", "Bill Amount", "Unit Cost"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Data Rows (Row 9-20)
    max_months = 12
    history1 = data_list[0].get("consumption_history", []) if len(data_list) > 0 else []
    history2 = data_list[1].get("consumption_history", []) if len(data_list) > 1 else []
    
    for i in range(max_months + 1): # Added one extra row as in the image
        row = 9 + i
        ws.cell(row=row, column=1, value=i+2).border = border # Sr.No starts from 2 in image
        
        # Consumer 1
        h1 = history1[i] if i < len(history1) else {}
        ws.cell(row=row, column=2, value=h1.get("month", "")).border = border
        ws.cell(row=row, column=3, value=h1.get("units", "")).border = border
        ws.cell(row=row, column=4, value=h1.get("bill_amount", "")).border = border
        ws.cell(row=row, column=5, value=h1.get("unit_cost", "")).border = border
        
        # Consumer 2
        h2 = history2[i] if i < len(history2) else {}
        ws.cell(row=row, column=6, value=h2.get("month", "")).border = border
        ws.cell(row=row, column=7, value=h2.get("units", "")).border = border
        ws.cell(row=row, column=8, value=h2.get("bill_amount", "")).border = border
        ws.cell(row=row, column=9, value=h2.get("unit_cost", "")).border = border

    # Calculations Row
    calc_start_row = 9 + max_months + 1
    
    def add_calc_row(row_idx, label, formula1, formula2=None, is_bold=True, fill1=None, fill2=None):
        ws.cell(row=row_idx, column=2, value=label).font = Font(bold=is_bold)
        ws.cell(row=row_idx, column=2).border = border
        
        c1 = ws.cell(row=row_idx, column=3, value=formula1)
        c1.border = border
        c1.font = Font(bold=is_bold)
        if fill1: c1.fill = fill1
        
        if formula2:
            c2 = ws.cell(row=row_idx, column=7, value=formula2)
            c2.border = border
            c2.font = Font(bold=is_bold)
            if fill2: c2.fill = fill2
        
        # Add empty borders for other columns
        for col in [1, 4, 5, 6, 8, 9]:
            ws.cell(row=row_idx, column=col).border = border

    # Average
    # Using IFERROR to handle empty columns
    add_calc_row(calc_start_row, "Average", f"=IFERROR(AVERAGE(C9:C21),0)", f"=IFERROR(AVERAGE(G9:G21),0)")
    ws.cell(row=calc_start_row, column=4, value=f"=IFERROR(AVERAGE(D9:D21),0)").border = border
    ws.cell(row=calc_start_row, column=5, value=f"=IFERROR(AVERAGE(E9:E21),0)").border = border
    ws.cell(row=calc_start_row, column=8, value=f"=IFERROR(AVERAGE(H9:H21),0)").border = border
    ws.cell(row=calc_start_row, column=9, value=f"=IFERROR(AVERAGE(I9:I21),0)").border = border

    # kW
    add_calc_row(calc_start_row + 1, "kW", f"=C{calc_start_row}/106.06", f"=G{calc_start_row}/106.06")
    
    # Solar Panels
    add_calc_row(calc_start_row + 2, "Solar Panels", f"=C{calc_start_row+1}/(B7/1000)", f"=G{calc_start_row+1}/(B7/1000)")
    
    # Solar capacity
    add_calc_row(calc_start_row + 3, "Solar capacity", f"=C{calc_start_row+4}*(B7/1000)", f"=G{calc_start_row+4}*(B7/1000)", fill1=yellow_fill, fill2=yellow_fill)
    ws.cell(row=calc_start_row + 3, column=2).fill = label_fill
    ws.cell(row=calc_start_row + 3, column=6).fill = label_fill
    
    # Number of Panels
    add_calc_row(calc_start_row + 4, "Number of Panels", f"=ROUNDUP(C{calc_start_row+2},0)", f"=ROUNDUP(G{calc_start_row+2},0)", fill1=green_fill, fill2=green_fill)
    ws.cell(row=calc_start_row + 4, column=2).fill = label_fill
    ws.cell(row=calc_start_row + 4, column=6).fill = label_fill

    # Final Totals
    ws.cell(row=calc_start_row + 6, column=2, value="Total solar capacity").font = bold_font
    ws.cell(row=calc_start_row + 6, column=3, value=f"=C{calc_start_row+3}+G{calc_start_row+3}").font = bold_font
    
    ws.cell(row=calc_start_row + 7, column=2, value="Number of solar panels").font = bold_font
    ws.cell(row=calc_start_row + 7, column=3, value=f"=C{calc_start_row+4}+G{calc_start_row+4}").font = bold_font

    # Final Totals
    ws.cell(row=calc_start_row + 6, column=2, value="Total solar capacity").font = bold_font
    ws.cell(row=calc_start_row + 6, column=3, value=f"=C{calc_start_row+3}+G{calc_start_row+3}").font = bold_font
    
    ws.cell(row=calc_start_row + 7, column=2, value="Number of solar panels").font = bold_font
    ws.cell(row=calc_start_row + 7, column=3, value=f"=C{calc_start_row+4}+G{calc_start_row+4}").font = bold_font

    # Save
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    # Test
    test_data = [
        {
            "consumer_name": "Shri Madhusham Khobragade",
            "consumer_number": "439320095567",
            "fixed_charges": 130,
            "sanctioned_load": "3.30KW",
            "connection_type": "90/ LT I Res 1-Phase",
            "consumption_history": [{"month": "Month "+str(i), "units": 100+i} for i in range(12)]
        }
    ]
    fill_excel_template(test_data, "test_output.xlsx")
